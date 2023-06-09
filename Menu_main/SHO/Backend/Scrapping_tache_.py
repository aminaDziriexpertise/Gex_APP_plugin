import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from unittest import result
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from flask import Flask
import re
import datetime;
import time
import string
from pyzwcad import ZwCAD, APoint
import array
import comtypes
import matplotlib as mpl
import matplotlib.pyplot as plt
from random import random
import numpy as np
import json
import requests
from flask import Flask,jsonify,request,make_response,url_for,redirect
# import load_workbook
from openpyxl import load_workbook


def test():
     acad = ZwCAD()
     url = requests.get('http://127.0.0.1:5000')
    
     print(url.status_code) 
     test = url.text
     # Définir un style de texte spécifique 
     acad.ActiveDocument.ActiveTextStyle = acad.ActiveDocument.TextStyles.Item("G100TXT-droit")
     #Scrapping tache déléguée
     driver = webdriver.Chrome() 
     #url = 'https://odoo-r7.gexpertise.fr/web/login'
     #url = 'https://odoo.gexpertise.fr/web#id=94933&action=311&active_id=18&model=helpdesk.ticket&view_type=form&menu_id=191'
     driver.get(test)
     username = driver.find_element(By.ID,"login")
     password = driver.find_element(By.ID,"password")
     #username.send_keys("i.benabdallah@gexpertise.fr")
     #password.send_keys("i$XnRR6H*l8d")

     username.send_keys("GexApp Pôle Développement")
     password.send_keys("Ppr9QX1RlqlB")

     # getting the button by class name 
     button1 = driver.find_element(By.XPATH,'//button[@type="submit"]').click()
     # clicking on the button
     #button1.click()
     
     def export(commune_name, address1, cadastre, Mission, num_livrable, Cree_le):
         # set file path
         filepath="test_model.xlsx"
         # load demo.xlsx 
         wb=load_workbook(filepath)
         sheet_ranges = wb['Titre']
         print(sheet_ranges['A1'].value)
         sheet_ranges['A1'] = commune_name
         sheet_ranges['A5'] = address1
         sheet_ranges['B8'] = cadastre
         sheet_ranges['C17'] = Mission
         sheet_ranges['C18'] = num_livrable
         sheet_ranges['G17'] = Cree_le
         sheet_ranges = wb['SDP']
         print(sheet_ranges['B11'].value)
         # save workbook
         wb.save(filepath)
     # function to create acronym
     def fxn(stng):
        # add first letter
        oupt = stng[0] 
        # iterate over string
        for i in range(1, len(stng)):
            if stng[i-1] == ' ':
                # add letter next to space
                oupt += stng[i]  
        # uppercase oupt
        oupt = oupt.upper()
        return oupt

     #**********************Nom_Mission **********************
     element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[5]/div[6]/table[2]/tbody/tr[5]/td[2]/a[@class='o_form_uri o_field_widget']")) 
        )
     Mission = element.get_attribute("innerHTML")
     Mission = Mission.split('-')
     Mission = Mission[0]
     #print(type(Mission))
     print('Mission :', Mission)
     

     #**********************Cree par **********************
     element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[5]/div[6]/table[2]/tbody/tr/td[2]/a[@class='o_form_uri o_field_widget field_partner_id']")) 
        )
     Cree_par = element.get_attribute("innerHTML")
     Cree_par = fxn(Cree_par)
     print('Cree_par :', Cree_par)
     print("tester la structure finale du nom", fxn(Cree_par))

    #**********************Full_address**********************
     element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[5]/div[6]/table[2]/tbody/tr[7]/td[2]/a[@class='o_form_uri o_field_widget o_readonly_modifier']")) 
        )
     full_address = element.get_attribute("innerHTML")

     driver.find_element(By.XPATH, ".//div[3]/div/div/div/div[5]/div[6]/table[2]/tbody/tr[7]/td[2]/a[@class='o_form_uri o_field_widget o_readonly_modifier']").click()
        
    #**********************Cree le **********************
     element1 = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[5]/div[6]/table/tbody/tr[8]/td[2]/span[@class='o_field_date o_field_widget o_readonly_modifier']")) 
        )
     global Cree_le
     Cree_le = element1.get_attribute("innerHTML")
     print('Cree_le :', Cree_le)
     Cree_le_extension = datetime.datetime.strptime(str(Cree_le),'%d/%m/%Y %H:%M:%S').strftime('%b %Y') #%B full name et %b short name
     print("forme_finale de la date ", Cree_le_extension)
     time.sleep(2) # Let the user actually see something!
     print("okyyyyyyyyy")
     #**********************long__lattt**********************
     element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[2]/div[3]/table[2]/tbody/tr/td[2]/span[@class='o_field_float o_field_number o_field_widget o_readonly_modifier']")) 
        )
     longitude = element.get_attribute("innerHTML")
     print("longitude", longitude)

     #***********************_lattt**********************
     element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[2]/div[3]/table[2]/tbody/tr[2]/td[2]/span[@class='o_field_float o_field_number o_field_widget o_readonly_modifier']")) 
        )
     lattitude = element.get_attribute("innerHTML")
     print("lattitude", lattitude)
     
      #***********************code lieu **********************
     element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//div[3]/div/div/div/div[2]/div[3]/table/tbody/tr[2]/td[2]/span[@class='o_field_char o_field_widget o_readonly_modifier']")) 
        )
     global num_livrable
     num_livrable = element.get_attribute("innerHTML")
     print("num_livrable", num_livrable)


     print('Full Address :', full_address)
     address = full_address.split(',')
     print("test_adress", address )
     address1 = address[0]
     print("L'ADRESSE est la suivante: ", address[0] )
    
     print("code postale est : ", address[1] )

     driver.quit()




     def localisation_extract(address):
           #Extraire 
            city_address = {'q': address}
            r = requests.get('https://geocodage.ign.fr/look4/address/search', params=city_address).json()
            #response= requests.get('https://geocodage.ign.fr/look4/address/search?q=140 boulevard Macdonald').json()
            adresse = r.get('features')[0].get('geometry').get('coordinates')
            city = r.get('features')[0].get('properties').get('city')
            print("city est ", city)
            a= "Paris 19e Arrondissement"
            if (a in city):
               city = " ".join(city.split()[1:])
               print("city_after modif", city )
            print("city_after modif", city )
           
            #Définition des polices pour des styles de texte spécifiques
            acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("swiss721CnBT.shx", False, False, 1, 0 or 0)

            t1 = acad.model.AddText("("+city+")", APoint(1653953.3941, 8210382.2722), 1)
            t1.color=0
            number = r.get('features')[0].get('properties').get('number')
            street = r.get('features')[0].get('properties').get('street')

            #Définition des polices pour des styles de texte spécifiques
            acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("swissl.shx", False, False, 1, 0 or 0)
            t1 = acad.model.AddText(number+", "+ street, APoint(1653953.2389, 8210374.8059), 1)
            t1.color=0
            t2 = acad.model.AddText(number+", "+ street, APoint(1653880.1227, 8210306.3645), 2)
            t2.color=0
            print(city, number, street)
            return (adresse)

   
     
     def cadastre_extract():
          n= localisation_extract(address1)
          filter = {'type':'point', 'coordinates': n}
          payload1 = {'searchGeom': json.dumps(filter)}
          #payload1 = {'searchGeom': json.dumps(coordinates)}
          #https://geocodage.ign.fr/look4/parcel/reverse?searchGeom={"type":"Point","coordinates":[2.379032,48.899039]}
          response= requests.get('https://geocodage.ign.fr/look4/parcel/reverse', params=payload1).json()
          #response= requests.get('https://geocodage.ign.fr/look4/parcel/reverse', params=payload1).json()
          print("ok123")
          global commune_name
          commune_name = response.get('features')[0].get('properties').get('nomCommune')
          if commune_name == "Paris":
             commune_name = "VILLE DE PARIS"
          print("commune name", commune_name)

          section_name = response.get('features')[0].get('properties').get('section')
          section_num  = response.get('features')[0].get('properties').get('numero')
          print(section_name)
     
       
          #Définition des polices pour des styles de texte spécifiques
          acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("suissb.shx", False, False, 1, 0 or 0)
          t1 = acad.model.AddText("Cadastre : Section "+ section_name + " n°  " + section_num,APoint(1653949.7439, 8210369.1671), 1)
          t1.color=0
          global cadastre
          cadastre = "Cadastre : Section "+section_name+ " n° "+ section_num
          
          #Définition des polices pour des styles de texte spécifiques
          acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("suissb.shx", False, False, 1, 0 or 0)
          t1 = acad.model.AddText(commune_name,APoint(1653955.3491, 8210384.8264), 1.5)
          t1.color=0
          

          #Autres cathégories de remplissage
          #Définition des polices pour des styles de texte spécifiques
          
          #------------------------Mission--------------------------------------
          acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("suissl.shx", False, False, 1, 0 or 0)
          t1 = acad.model.AddText("Mission : " + Mission + " -Lieu :",APoint(1653954.6317, 8210310.2474), 0.7)
          t1.color=0

          #-------------------------Code Lieu --------------------------
          acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("suissl.shx", False, False, 1, 0 or 0)
          t1 = acad.model.AddText(num_livrable + " - Indice : ",APoint(1653956.3518, 8210308.9432), 0.7)
          t1.color=0
        
       

          #-------------------------Date et responsables --------------------------
          acad.ActiveDocument.TextStyles.Item("G100TXT-droit").SetFont("suissl.shx", False, False, 1, 0 or 0)
          t1 = acad.model.AddText("Date : "+ Cree_le +" - Resp : "+ Cree_par,APoint(1653953.5747, 8210308.0091), 0.7)
          t1.color=0
          
          export (commune_name, address1, cadastre, Mission, num_livrable, Cree_le)

     cadastre_extract()
  


    


    
if __name__ == "__main__":
   test()

